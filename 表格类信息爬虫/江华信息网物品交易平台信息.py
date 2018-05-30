# 存储信息到excel表中
# 进一步完善，名称，发布时间，价格，区域，联系人，联系电话
import os
import re
import requests
import urllib.request
from bs4 import BeautifulSoup
from PIL import Image
import pytesseract
from openpyxl import Workbook


headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 \
    (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36',
    'Host': 'info.425500.cn',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1'
}


def img_to_str(image_path):
    # 识别图片上的电话
    return pytesseract.image_to_string(Image.open(image_path))


def save_to_excel(total_item_info):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "交易信息"
    sheet.cell(row=1, column=1, value='标题')
    sheet.cell(row=1, column=2, value='发布时间')
    sheet.cell(row=1, column=3, value='浏览次数')
    sheet.cell(row=1, column=4, value='所在区域')
    sheet.cell(row=1, column=5, value='联系电话')
    sheet.cell(row=1, column=6, value='交易信息详细链接')

    for i in range(2, len(total_item_info)+2):
        try:
            sheet.cell(row=i, column=1, value=total_item_info[i - 2]['title'])
            sheet.cell(row=i, column=2, value=total_item_info[i - 2]['is_overdue'])
            sheet.cell(row=i, column=3, value=total_item_info[i - 2]['view_num'])
            sheet.cell(row=i, column=4, value=total_item_info[i - 2]['area'])
            sheet.cell(row=i, column=5, value=total_item_info[i - 2]['tel'])
            sheet.cell(row=i, column=6, value=('=HYPERLINK("%s")' % \
                                               (total_item_info[i - 2]['link'])))
        except:
            print("数据写入出错")
    wb.save("江华信息网物品交易信息.xlsx")  # EXCEL保存


def get_title_area_tel(link):
    info_page = requests.get(link, headers=headers).content
    soup = BeautifulSoup(info_page, 'lxml')
    title = soup.select_one("h1.news-title").get_text()  # 交易信息标题
    is_overdue = soup.find('span', class_="news-span1").get_text()  # 交易信息是否过期
    info_id = re.findall('\d+', link.split('/')[-1])[0]  # 通过网址获取交易编号
    # 交易信息浏览次数是通过后台脚本处理的:http://info.425500.cn/public/ajax.aspx?action=addnum&id=交易编号&t=1
    info_id_add_url = 'http://info.425500.cn/public/ajax.aspx?action=addnum&id=%s&t=1' % info_id
    view_num = requests.get(info_id_add_url).text
    view_num = "浏览了" + re.findall('\d+', view_num)[0] + "次"   # 正则取出交易浏览次数
    area = soup.select("div.newscontent2 > div.nc2-content > ul > li > span.cBlack")[0].get_text().lstrip().rstrip()
    tel_img_down_path = './telImg/'
    if not os.path.exists(tel_img_down_path):
        os.mkdir(tel_img_down_path)
    try:
        # 部分联系电话仅限会员查看，需要登录，这里检测如果需要登录，就返回空信息，不获取需要登录才能查看的电话
        tel_img_src = soup.select("div.newscontent2 > div.nc2-content > ul > li > span.cBlack > img")[0].get("src")
        tel_img_src_full = "http://info.425500.cn" + tel_img_src[2:]
        tel_img_title = title + '.png'
        # 将电话图片下载到指定文件夹，用tesseract识别出电话号码
        urllib.request.urlretrieve(tel_img_src_full,
                                   os.path.join(tel_img_down_path, tel_img_title))
        tel = img_to_str(tel_img_down_path + tel_img_title)
        info_dict = {
            'title': title,
            'area': area,
            'tel': tel,
            'is_overdue': is_overdue,
            'view_num': view_num,
            'link': link
        }
        return info_dict
    except:
        print("Error: " + title)
        return None


def get_all_links(total_item_info, root_url, page_id):
    url = root_url + "category-1-0-0-0-p" + str(page_id) + ".html"
    print("爬取目标页面：" + url)
    items_page = requests.get(url, headers=headers).content
    soup = BeautifulSoup(items_page, 'lxml')
    item_links = soup.select('a.list2li-a1')
    for item_link in item_links:
        link = item_link['href']
        item_info_dict = get_title_area_tel(link)
        if item_info_dict:
            total_item_info.append(item_info_dict)
    print("目标页面信息获取成功")


def get_total_page_num(root_url):
    first_page_url = root_url + 'category-1-0-0-0-p1.html'
    first_page = requests.get(first_page_url, headers=headers).content
    soup = BeautifulSoup(first_page, 'lxml')
    page_num = soup.select_one('div.AntPage > ul > li > span.total').get_text()
    page_num = page_num.split('/')[1].split(' ')[0]
    return page_num


if __name__ == '__main__':
    root_url = 'http://info.425500.cn/'
    # total_page_num = get_total_page_num(root_url)
    # 将total_page_num改为 get_total_page_num(root_url)返回值可以获取全部数据
    total_page_num = 3
    total_item_info = []
    for page_id in range(1, total_page_num+1):
        get_all_links(total_item_info, root_url, page_id)
    save_to_excel(total_item_info)
    print("爬取江华信息网物品交易平台所有数据完成，信息存储在当前目录下的江华信息网物品交易信息.xlsx文件中")



