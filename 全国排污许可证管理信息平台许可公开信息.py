# coding=utf-8

import requests
import re
from bs4 import BeautifulSoup
from openpyxl import *

headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 \
    (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36',
    'Host': 'permit.mep.gov.cn',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1'
}


def crawler_pageform(root_url, end_page_num):
    """读取某页表格数据"""
    data = []
    # 翻页
    for page_num in range(1, end_page_num+1):
        page_num = str(page_num)
        page_url = root_url + '?pageno=%s' % (page_num)
        req = requests.get(page_url, headers=headers)
        if req.status_code == 200:
            print("获取第%s页成功,正在读取该页数据" % page_num)
            soup = BeautifulSoup(req.content, 'lxml')
            try:
                trs = soup.find('div', class_='tb-con').find_all('tr')[1:]
                # 存储该页表格数据
                page_data = []
                for tr in trs:
                    row_data = []
                    for td in tr:
                        # 过滤掉'\n'字符和最后一个空的链接内容
                        if td.string != '\n' and td.string:
                            row_data.append(td.string)
                    link = 'http://permit.mep.gov.cn' + \
                           tr.select('td.bgcolor1 > a')[0]['href']
                    row_data.append(link)
                    page_data.append(row_data)
                data += page_data
                print("读取第%s页数据成功" % page_num)

            except Exception as e:
                print("读取第%s页数据失败" % page_num)
                print(e)
        else:
            print("获取第%s页失败" % page_num)
    return data


def write_to_excel(data):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "许可信息公开数据"
    sheet.cell(row=1, column=1, value='编号')
    sheet.cell(row=1, column=2, value='省/直辖市')
    sheet.cell(row=1, column=3, value='地市')
    sheet.cell(row=1, column=4, value='许可证编号')
    sheet.cell(row=1, column=5, value='单位名称')
    sheet.cell(row=1, column=6, value='行业类别')
    sheet.cell(row=1, column=7, value='有效期限')
    sheet.cell(row=1, column=8, value='发证日期')
    sheet.cell(row=1, column=9, value='查看链接')

    for i in range(2, len(data) + 2):
        # 编号
        sheet.cell(row=i, column=1, value=i-1)
        for j in range(1, 8):
            try:
                sheet.cell(row=i, column=j+1, value=data[i-2][j-1])
                # 最后一个单元格链接单独处理
                sheet.cell(row=i, column=9, value=('=HYPERLINK("%s")' % (data[i-2][7])))
            except Exception as e:
                print("第" + i + "行", "第" + j + "列数据写入出错")
                print(e)
    wb.save("全国排污许可证管理信息平台许可公开信息.xlsx")  # EXCEL保存


def get_end_page_num(root_url):
    end_page_data = requests.get(root_url, headers=headers).content
    soup = BeautifulSoup(end_page_data, 'lxml')
    end_page_text = soup.find('div', class_='page').get_text()
    # 获取最后一页页码
    end_page_num = re.findall('\d+', end_page_text)[0]
    return end_page_num


if __name__ == '__main__':
    root_url = 'http://permit.mep.gov.cn/permitExt/outside/Publicity'
    # end_page_num = get_end_page_num(root_url)
    # 将4换为end_page_num可以获取全部数据
    end_page_num = 4
    data = crawler_pageform(root_url, end_page_num)
    write_to_excel(data)
    print("爬取全国排污许可证管理信息平台许可公开信息所有数据完成，信息存储在当前目录下的全国排污许可证管理信息平台许可公开信息.xlsx文件中")












